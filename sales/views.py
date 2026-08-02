from django.shortcuts import render, redirect, get_object_or_404
from products.models import Product
from .models import Sale
from customers.models import Customer
from django.db.models import Sum, Q
from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
import json
from django.db.models.functions import TruncMonth
def sales_page(request):

    products = Product.objects.all()
    customers = Customer.objects.all()

    if request.method == "POST":

        customer_id = request.POST["customer"]
        product_id = request.POST["product"]
        quantity = int(request.POST["quantity"])


        customer = Customer.objects.get(id=customer_id)
        product = Product.objects.get(id=product_id)


        if quantity <= product.quantity:


            total = product.price * quantity


            profit = (
                product.price - product.cost_price
            ) * quantity


            sale = Sale.objects.create(

                customer=customer,

                product=product,

                quantity=quantity,

                total_price=total,

                profit=profit

            )


            product.quantity -= quantity

            product.save()


            return redirect(
                "receipt",
                sale_id=sale.id
            )


        else:

            return render(

                request,

                "sales.html",

                {
                    "products": products,

                    "customers": customers,

                    "error": "Not enough stock."

                }

            )


    return render(

        request,

        "sales.html",

        {
            "products": products,

            "customers": customers

        }

    )

def sales_history(request):

    search = request.GET.get("search", "")

    sales = Sale.objects.all().order_by("-sale_date")

    if search:

        sales = sales.filter(
            Q(product__product_name__icontains=search) |
            Q(customer__full_name__icontains=search)
        )

    return render(
        request,
        "sales_history.html",
        {
            "sales": sales,
            "search": search
        }
    )

def receipt(request, sale_id):
    sale = get_object_or_404(Sale, id=sale_id)

    return render(
        request,
        "receipt.html",
        {
            "sale": sale
        }
    )


def reports(request):

    sales = Sale.objects.select_related(
        "product",
        "customer"
    ).order_by("-sale_date")


    # Date filter

    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")


    if from_date:
        sales = sales.filter(
            sale_date__date__gte=from_date
        )


    if to_date:
        sales = sales.filter(
            sale_date__date__lte=to_date
        )


    # Statistics

    total_products = Product.objects.count()

    total_customers = Customer.objects.count()

    total_sales = sales.count()


    total_revenue = (
        sales.aggregate(
            revenue=Sum("total_price")
        )["revenue"] or 0
    )


    total_profit = (
        sales.aggregate(
            profit=Sum("profit")
        )["profit"] or 0
    )


    low_stock = Product.objects.filter(
        quantity__lt=5
    )


    # -------------------------
    # Product Chart
    # -------------------------

    chart_labels = []
    chart_values = []


    product_sales = (
        sales.values(
            "product__product_name"
        )
        .annotate(
            total_qty=Sum("quantity")
        )
    )


    for item in product_sales:

        chart_labels.append(
            item["product__product_name"]
        )

        chart_values.append(
            item["total_qty"]
        )


    # Pie chart data

    pie_labels = chart_labels

    pie_values = chart_values



    # -------------------------
    # Monthly Revenue
    # -------------------------

    month_labels = []
    month_values = []


    monthly = (
        sales.annotate(
            month=TruncMonth("sale_date")
        )
        .values("month")
        .annotate(
            revenue=Sum("total_price")
        )
        .order_by("month")
    )


    for item in monthly:

        if item["month"]:

            month_labels.append(
                item["month"].strftime("%b %Y")
            )


            month_values.append(
                float(item["revenue"])
            )



    context = {


        "sales": sales,


        "from_date": from_date,

        "to_date": to_date,


        "total_products": total_products,

        "total_customers": total_customers,

        "total_sales": total_sales,

        "total_revenue": total_revenue,

        "total_profit": total_profit,


        "low_stock": low_stock,


        "chart_labels": json.dumps(chart_labels),

        "chart_values": json.dumps(chart_values),



        "pie_labels": json.dumps(pie_labels),

        "pie_values": json.dumps(pie_values),



        "month_labels": json.dumps(month_labels),

        "month_values": json.dumps(month_values),

    }


    return render(
        request,
        "reports.html",
        context
    )
def sales_excel(request):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Sales Report"


    sheet["A1"] = "POS MANAGEMENT SYSTEM"
    sheet["A1"].font = Font(bold=True, size=16)

    sheet["A2"] = "Sales Report"
    sheet["A2"].font = Font(bold=True, size=14)


    headings = [
        "Receipt",
        "Customer",
        "Product",
        "Quantity",
        "Total Price",
        "Profit",
        "Date"
    ]


    row = 4

    for col, heading in enumerate(headings, start=1):

        cell = sheet.cell(
            row=row,
            column=col
        )

        cell.value = heading
        cell.font = Font(bold=True)



    sales = Sale.objects.select_related(
        "customer",
        "product"
    )


    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")


    if from_date and from_date != "None":
        sales = sales.filter(
            sale_date__date__gte=from_date
        )


    if to_date and to_date != "None":
        sales = sales.filter(
            sale_date__date__lte=to_date
        )


    row = 5

    total_revenue = 0
    total_profit = 0


    for sale in sales:


        sheet.cell(row=row, column=1).value = sale.id


        sheet.cell(row=row, column=2).value = (
            sale.customer.full_name
            if sale.customer
            else "Walk-in Customer"
        )


        sheet.cell(row=row, column=3).value = (
            sale.product.product_name
        )


        sheet.cell(row=row, column=4).value = sale.quantity


        sheet.cell(row=row, column=5).value = float(
            sale.total_price
        )


        sheet.cell(row=row, column=6).value = float(
            sale.profit
        )


        sheet.cell(row=row, column=7).value = (
            sale.sale_date.strftime("%d/%m/%Y")
        )


        total_revenue += sale.total_price

        total_profit += sale.profit


        row += 1



    sheet.cell(row=row+1, column=4).value = "Total Revenue"

    sheet.cell(row=row+1, column=5).value = float(
        total_revenue
    )


    sheet.cell(row=row+2, column=4).value = "Total Profit"

    sheet.cell(row=row+2, column=5).value = float(
        total_profit
    )



    response = HttpResponse(
        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


    response["Content-Disposition"] = (
        'attachment; filename="Sales_Report.xlsx"'
    )


    workbook.save(response)


    return response

def sales_pdf(request):

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="Sales_Report.pdf"'

    doc = SimpleDocTemplate(response)

    styles = getSampleStyleSheet()

    elements = []

    elements.append(Paragraph("<b>POS MANAGEMENT SYSTEM</b>", styles["Title"]))
    elements.append(Paragraph("Sales Report", styles["Heading2"]))

    data = [
        ["Receipt", "Customer", "Product", "Qty", "Total", "profit", "Date"]
    ]

    sales = Sale.objects.select_related(
        "customer",
        "product"
    )

    # Filter by date
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")

    if from_date and from_date != "None":
     sales = sales.filter(
        sale_date__date__gte=from_date
    )

    if to_date and to_date != "None":
     sales = sales.filter(
        sale_date__date__lte=to_date
    )

    total_revenue = 0

    for sale in sales:

        data.append([
            sale.id,
            sale.customer.full_name if sale.customer else "Walk-in Customer",
            sale.product.product_name,
            sale.quantity,
            f"UGX {sale.total_price}",
            f"UGX {sale.profit}",
            sale.sale_date.strftime("%d/%m/%Y")
        ])

        total_revenue += sale.total_price

    data.append(
    ["", "", "", "Total Revenue", f"UGX {total_revenue}", "", ""]
   )

    table = Table(data)

    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.darkblue),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 1, colors.black),
        ("BACKGROUND", (0,1), (-1,-2), colors.beige),
        ("BACKGROUND", (0,-1), (-1,-1), colors.lightgrey),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
    ]))

    elements.append(table)

    doc.build(elements)

    return response